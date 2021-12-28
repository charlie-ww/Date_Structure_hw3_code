import random
from openpyxl import Workbook
import numpy as np
import time

# Hash Table
class Node:
    def __init__(self, key=None, data=None):
        self.value = {}
        self.value[key] = data
        self.next = None

    def __repr__(self):
        return str(self.data)


class LinkedList:
    def __init__(self, head=None):
        self.head = head
        self.next = None
        self.count = 0

    def __str__(self):
        str_list = []
        current = self.head
        while current:
            str_list.append(str(current.value))
            current = current.next
        return "[" + "->".join(str_list) + "]"

    def __repr__(self):
        return str(self)


class HashTable:
    def __init__(self, size):
        self.size = size
        self.values = [None] * size
        self.length = 0

    def hash(self, key, size):
        hashCode = 0
        for i in range(len(key)):
            hashCode += ord(key[i])
        return hashCode % size

    def add(self, key, value):
        hashIndex = self.hash(key, self.size)
        node = Node(key, value)
        if not self.values[hashIndex]:
            self.values[hashIndex] = LinkedList(node)
        else:
            current = self.values[hashIndex].head
            while current.next:
                current = current.next
            current.next = node
        self.values[hashIndex].count += 1
        self.length += 1

    def search(self, key):
        hashIndex = self.hash(key, self.size)
        slot = self.values[hashIndex]
        current = slot.head
        if key in current.value:
            return current.value
        while current.next:
            if key in current.next.value:
                return current.next.value
            else:
                current = current.next
        return "Data not found"

    def remove(self, key):
        hashIndex = self.hash(key, self.size)
        slot = self.values[hashIndex]
        current = slot.head
        if key in current.value:
            current = current.next
            slot.count -= 1
            self.length -= 1
            return "Data was deleted successfully"
        while current.next:
            if key in current.next.value:
                current.next = current.next.next
                slot.count -= 1
                self.length -= 1
                return "Data was deleted successfully"
            else:
                current = current.next
        return "Data is not exhausting"

    def __repr__(self):
        return str(self.values)

wb = Workbook()
ws = wb.active
ws['A1'] = '時間'
ws['B1'] = '插入'
ws['C1'] = '搜尋'

for a in range(10,20):
    N = 2 ** 30 + 1
    size = 2 ** a
    ht = HashTable(N)
    random_list = np.random.randint(0, N, size=size)
    string_ints = [str(int) for int in random_list]
    str_of_ints = ",".join(string_ints)
    start = time.time()
    for i in range(size):
        ht.add(str_of_ints[i], str_of_ints[i])
    end = time.time()
    print("use ", end-start, "seconds")

    randomsearch_list = np.random.randint(0, N, size=100000)
    stringsearch_ints = [str(int) for int in randomsearch_list]
    strsearch_of_ints = ",".join(stringsearch_ints)
    start1 = time.time()
    for i in range(100000):
        ht.search(strsearch_of_ints[i])
    end1 = time.time()
    print("use ", end1-start1, "seconds")
    ws.append([a, end-start, end1-start1])
    wb.save("hashtable.xlsx")
